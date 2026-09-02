#!/usr/bin/env python3
"""Импорт распределения по когортам из таблицы ВШМ в данные бота.

    python scripts/import_cohorts.py Cohorts_Distribution_MiM_2026.xlsx

Читает лист с колонками «No.», «Name, Last Name» и по одной колонке на
предмет, после чего складывает результат в bot/roster/mim_2026.json.
Скрипт нужен раз в год, когда деканат публикует новое распределение,
поэтому openpyxl вынесен в requirements-dev.txt и боту не нужен.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

DEFAULT_OUTPUT = ROOT / "bot" / "roster" / "mim_2026.json"

# Как колонка таблицы называется в расписании на сайте. `match` — подстроки,
# по которым занятие относят к предмету (в нижнем регистре, латиница и
# кириллица). `kind`: cohort — метка вида «Coh.1», group — просто номер,
# educator — распределение по преподавателю.
SUBJECTS = [
    {
        "key": "corp_finance",
        "column": "Corp. Finance",
        "title": "Corporate Finance",
        "kind": "cohort",
        "match": ["corporate finance", "corp. finance", "corp finance", "корпоративные финансы"],
    },
    {
        "key": "org_behaviour",
        "column": "Org. Behaviour",
        "title": "Organizational Behaviour",
        "kind": "cohort",
        "match": [
            "organizational behaviour",
            "organizational behavior",
            "org. behaviour",
            "org behaviour",
            "организационное поведение",
        ],
    },
    {
        "key": "qmbr_lectures",
        "column": "QMBR lectures",
        "title": "QMBR, лекции",
        "kind": "cohort",
        "match": ["qmbr", "quantitative methods", "количественные методы"],
        "event_type": "lecture",
    },
    {
        "key": "qmbr_seminars",
        "column": "QMBR seminars",
        "title": "QMBR, семинары",
        "kind": "cohort",
        "match": ["qmbr", "quantitative methods", "количественные методы"],
        "event_type": "seminar",
    },
    {
        "key": "rs_1",
        "column": "RS I",
        "title": "Research Seminar I",
        "kind": "group",
        "match": ["research seminar", "rs i", "исследовательский семинар"],
    },
    {
        "key": "ccm",
        "column": "CCM",
        "title": "Cross-Cultural Management",
        "kind": "group",
        "match": [
            "cross-cultural management",
            "cross cultural management",
            "ccm",
            "кросс-культурный менеджмент",
        ],
    },
    {
        "key": "mps_1",
        "column": "MPS I",
        "title": "MPS I",
        "kind": "educator",
        "match": ["mps", "managerial problem solving", "управленческ"],
    },
]


def normalize_cohort(value: object) -> str:
    """«Coh. 2» и «Coh.2» — одно и то же; номер группы приводим к строке."""
    text = " ".join(str(value or "").split())
    match = re.fullmatch(r"(?i)coh\.?\s*(\d+)", text)
    if match:
        return f"Coh.{match.group(1)}"
    return text


def read_students(path: Path, sheet: str | None) -> list[dict]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook[sheet] if sheet else workbook[workbook.sheetnames[0]]

    rows = list(worksheet.iter_rows(values_only=True))
    header = [" ".join(str(cell or "").split()) for cell in rows[0]]
    index = {name: position for position, name in enumerate(header)}

    missing = [item["column"] for item in SUBJECTS if item["column"] not in index]
    if missing:
        raise SystemExit(
            f"В таблице нет колонок: {', '.join(missing)}.\n"
            f"Найдены: {', '.join(filter(None, header))}"
        )

    name_column = next(
        (position for name, position in index.items() if "name" in name.lower()), 1
    )

    number_column = next(
        (position for name, position in index.items() if name.lower().startswith("no")), 0
    )

    students: list[dict] = []
    for row in rows[1:]:
        full_name = " ".join(str(row[name_column] or "").split())
        if not full_name:
            continue
        parts = full_name.split()
        students.append(
            {
                "no": str(row[number_column] or "").strip(),
                "name": full_name,
                # В таблице фамилия идёт первой; у иностранных студентов
                # порядок другой, поэтому в поиске участвуют все части имени.
                "last_name": parts[0],
                "parts": parts,
                "cohorts": {
                    item["key"]: normalize_cohort(row[index[item["column"]]])
                    for item in SUBJECTS
                },
            }
        )
    return students


def deduplicate(students: list[dict]) -> tuple[list[dict], list[str], list[str]]:
    """Убирает строки-копии и находит настоящих однофамильцев.

    В таблице встречаются повторы одной и той же строки. Если у повторов
    совпадают и имя, и все когорты, лишнюю запись можно выбросить без
    потерь. Если же у тёзок когорты разные — это разные люди, их оставляем,
    а бот при выборе покажет номер из таблицы.
    """
    unique: list[dict] = []
    seen: dict[tuple, dict] = {}
    removed: list[str] = []
    for student in students:
        signature = (student["name"], tuple(sorted(student["cohorts"].items())))
        if signature in seen:
            removed.append(f"№{student['no']} {student['name']}")
            continue
        seen[signature] = student
        unique.append(student)

    by_name: dict[str, int] = {}
    for student in unique:
        by_name[student["name"]] = by_name.get(student["name"], 0) + 1
    namesakes = sorted(name for name, count in by_name.items() if count > 1)
    return unique, removed, namesakes


def main() -> None:
    parser = argparse.ArgumentParser(description="Импорт когорт MiM из xlsx")
    parser.add_argument("xlsx", type=Path, help="файл распределения от деканата")
    parser.add_argument("--sheet", help="имя листа (по умолчанию первый)")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--program", default="MiM 2026", help="подпись программы")
    args = parser.parse_args()

    students = read_students(args.xlsx, args.sheet)
    students, removed, namesakes = deduplicate(students)
    payload = {
        "program": args.program,
        "source": args.xlsx.name,
        "subjects": SUBJECTS,
        "students": students,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, ensure_ascii=False, indent=1) + "\n", encoding="utf-8"
    )

    print(f"Студентов: {len(students)}")
    if removed:
        print(f"Убраны строки-копии ({len(removed)}): {', '.join(removed)}")
    if namesakes:
        print(f"Полные тёзки с разными когортами: {', '.join(namesakes)}")
    for item in SUBJECTS:
        values = sorted({student["cohorts"][item["key"]] for student in students})
        print(f"  {item['column']:<16} {', '.join(values)}")
    print(f"Записано: {args.output.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
